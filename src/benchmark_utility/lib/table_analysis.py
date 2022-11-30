#!/usr/bin/python
import sys,os
# sys.path.append('../')
sys.path.insert(0, os.getcwd())
from src.amr_utility import name_utility,file_utility
import json
import pandas as pd
import numpy as np
import copy
from openpyxl import load_workbook
from src.benchmark_utility.lib.CombineResults import combine_data
import itertools
from scipy.stats import ttest_rel

'''
This script organizes the performance for Supplementary materials, and further analysis on the results.
'''


def extract_info(level,s,fscore, f_all,output_path,step,tool_list,foldset,com_tool_list):
    main_meta,_=name_utility.GETname_main_meta(level)
    data = pd.read_csv(main_meta, index_col=0, dtype={'genome_id': object}, sep="\t")
    data = data[data['number'] != 0]  # drop the species with 0 in column 'number'.
    if f_all == False:#should not use it.
        data = data.loc[s, :]
    species_list=['Escherichia coli','Staphylococcus aureus','Salmonella enterica','Klebsiella pneumoniae','Pseudomonas aeruginosa',
                  'Acinetobacter baumannii','Streptococcus pneumoniae','Mycobacterium tuberculosis', 'Campylobacter jejuni',
                  'Enterococcus faecium','Neisseria gonorrhoeae']
    data=data.loc[species_list,:]
    df_species = data.index.tolist()
    antibiotics= data['modelling antibiotics'].tolist()

    # ------------------------------------------
    # Step 1.1 figuring out ML generally performance. and compare with ResFinder
    # ------------------------------------------
    # ------------------------------------------
    # # Step 1.2 figuring out the number of combinations that each tool performs better than ML baseline
    # # ------------------------------------------
    if step=='1':
        if com_tool_list==['Point-/ResFinder']:
            path_table_results2=output_path+ 'Results/other_figures_tables/ML_Com_resfinder_'+fscore+'.xlsx'
        elif com_tool_list==['ML Baseline (Majority)']:
            path_table_results2=output_path+ 'Results/other_figures_tables/ML_Com_MLbaseline_'+fscore+'.xlsx'
        else:
            print('Please add a new name manually at ./src/benchmark_utility/lib/table_analysis.py \
            if a new software except resfinder and ML baseline is chosed for a comparison with the rest.')
            exit(1)


        df1 = pd.DataFrame(index=species_list)
        file_utility.make_dir(os.path.dirname(path_table_results2))
        df1.to_excel(path_table_results2, sheet_name='introduction')
        score_list=['f1_macro', 'f1_positive', 'f1_negative', 'accuracy','clinical_f1_negative','clinical_precision_neg', 'clinical_recall_neg']
        for com_tool in com_tool_list:
            #each time count the cases the com_tool outperforms others.
            for eachfold in foldset:
                df_final=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software']+score_list)
                df_com=pd.DataFrame(columns=['species', 'antibiotics', 'folds', 'software']+score_list)
                print('Compare with:',com_tool,eachfold,'-----')
                for species, antibiotics_selected in zip(df_species, antibiotics):
                    species_sub=[species]
                    tool_list_rest=copy.deepcopy(tool_list)
                    df_macro=combine_data(species_sub,level,'f1_macro',tool_list_rest,[eachfold],output_path)
                    df_acu=combine_data(species_sub,level,'accuracy',tool_list_rest,[eachfold],output_path)
                    df_neg=combine_data(species_sub,level,'f1_negative',tool_list_rest,[eachfold],output_path)
                    df_pos=combine_data(species_sub,level,'f1_positive',tool_list_rest,[eachfold],output_path)
                    df_cl_f1=combine_data(species_sub,level,'clinical_f1_negative',tool_list_rest,[eachfold],output_path)
                    df_cl_pre=combine_data(species_sub,level,'clinical_precision_neg',tool_list_rest,[eachfold],output_path)
                    df_cl_rec=combine_data(species_sub,level,'clinical_recall_neg',tool_list_rest,[eachfold],output_path)

                    df_macro['f1_negative']=df_neg['f1_negative']
                    df_macro['f1_positive']=df_pos['f1_positive']
                    df_macro['accuracy']=df_acu['accuracy']
                    df_macro['clinical_f1_negative']=df_cl_f1['clinical_f1_negative']
                    df_macro['clinical_precision_neg']=df_cl_pre['clinical_precision_neg']
                    df_macro['clinical_recall_neg']=df_cl_rec['clinical_recall_neg']


                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])
                    df_macro = df_macro[['species', 'antibiotics', 'folds', 'software']+score_list]
                    df_final= pd.concat([df_final,df_macro])

                    # -----compare tool, based on fscore scores.
                    df_com_sub=combine_data(species_sub,level,fscore,[com_tool],[eachfold],output_path)
                    df_com= pd.concat([df_com,df_com_sub])


                df_com[fscore] = df_com[fscore] .astype(str)
                #### df_com['compare_'+fscore]=df_com[fscore].apply(lambda x:x.split('±')[0] if (len(x.split('±'))==2 ) else x)
                df_com['compare_'+fscore+'_mean']=df_com[fscore].apply(lambda x:x.split('±')[0])
                df_com['compare_'+fscore+'_std']=df_com[fscore].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else np.nan)
                df_com=df_com[['species', 'antibiotics','compare_'+fscore+'_mean','compare_'+fscore+'_std']]

                df_final[fscore] = df_final[fscore] .astype(str)

                df_final[fscore+'_mean']=df_final[fscore].apply(lambda x:x.split('±')[0])
                df_final[fscore+'_std']=df_final[fscore].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else np.nan)
                # -----------------------------------
                # ---make a comparison

                df_final=pd.merge(df_final, df_com, how="left", on=['species', 'antibiotics'])

                # ----------------------------------------


                df_final[fscore+'_mean'] = df_final[fscore+'_mean'] .astype(float)
                df_final[fscore+'_std'] = df_final[fscore+'_std'] .astype(float)
                df_final['compare_'+fscore+'_mean'] = df_final['compare_'+fscore+'_mean'] .astype(float)
                df_final['compare_'+fscore+'_std'] = df_final['compare_'+fscore+'_std'] .astype(float)

                wb = load_workbook(path_table_results2)
                ew = pd.ExcelWriter(path_table_results2)
                ew.book = wb
                df_final.to_excel(ew,sheet_name = (eachfold.split(' ')[0][0]+eachfold.split(' ')[1][0]+'_Comp_'+str(com_tool.translate(str.maketrans({'/': '', ' ': '_'})))))
                ew.save()



    # ------------------------------------------
    # Step 2 figuring out which ML performs best.
    # ------------------------------------------
    if step=='2':
        if tool_list==['Point-/ResFinder', 'Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']:

            if fscore=='f1_macro':
                path_table_results3_1=output_path+ 'Results/supplement_figures_tables/S6-1_software_winner_'+fscore+'.xlsx'
                path_table_results3_2=output_path+ 'Results/final_figures_tables/F3_results_heatmap_'+fscore+'.xlsx'
            else: #clinical-oriented
                path_table_results3_1=output_path+ 'Results/other_figures_tables/software_winner_'+fscore+'.xlsx'
                path_table_results3_2=output_path+ 'Results/other_figures_tables/results_heatmap_'+fscore+'.xlsx'
        elif tool_list==['Point-/ResFinder', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover','Single-species-antibiotic Aytan-Aktug',
                   'Single-species multi-antibiotics Aytan-Aktug','Discrete databases multi-species model',
                'Concatenated databases mixed multi-species model', 'Concatenated databases leave-one-out multi-species model']:
            if fscore=='f1_macro':
                path_table_results3_1=output_path+ 'Results/supplement_figures_tables/S6-2_software_winner_multiModel_'+fscore+'.xlsx'
                path_table_results3_2=output_path+ 'Results/final_figures_tables/results_heatmap_multiModel_'+fscore+'.xlsx'
            else:#clinical-oriented
                path_table_results3_1=output_path+ 'Results/other_figures_tables/software_winner_multiModel_'+fscore+'.xlsx'
                path_table_results3_2=output_path+ 'Results/other_figures_tables/results_heatmap_multiModel_'+fscore+'.xlsx'
        else:
            print('Please add a new name manually at ./src/benchmark_utility/lib/table_analysis.py \
            if new software combinations are used for deciding winner or generate heatmap format excel.')
            exit(1)

        df1 = pd.DataFrame(index=species_list)
        file_utility.make_dir(os.path.dirname(path_table_results3_1))
        df1.to_excel(path_table_results3_1, sheet_name='introduction')

        #### each time count the cases the com_tool outperforms others.
        for eachfold in foldset:

            i=0
            for each_tool in tool_list:
                print(each_tool,eachfold,'-----')
                df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                for species, antibiotics_selected in zip(df_species, antibiotics):

                    species_sub=[species]
                    df_macro=combine_data(species_sub,level,fscore,[each_tool],[eachfold],output_path)
                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])

                    df_macro[fscore] = df_macro[fscore].astype(str)
                    df_macro[each_tool]=df_macro[fscore].apply(lambda x:x.split('±')[0])
                    df_macro[each_tool+'_std']=df_macro[fscore].apply(lambda x: x.split('±')[1] if (len(x.split('±'))==2) else 10)
                    df_macro[each_tool] = df_macro[each_tool] .astype(float)
                    df_macro[each_tool+'_std'] = df_macro[each_tool+'_std'] .astype(float)

                    df_macro=df_macro[['species', 'antibiotics',each_tool,each_tool+'_std']]
                    df_final= pd.concat([df_final,df_macro])

                if i==0:
                    df_compare=df_final
                else:
                    df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
                i+=1




            df_std=df_compare[[x+'_std' for x in tool_list]]

            df_compare['max_'+fscore]=df_compare[tool_list].max(axis=1)
            a = df_compare[tool_list]
            df = a.eq(a.max(axis=1), axis=0)

            #considering of std
            for index, row in df.iterrows():

                winner=[]
                winner_std=[]
                for columnIndex, value in row.items():
                    # print(columnIndex,value, end="\t")
                    if value==True:
                        winner.append(columnIndex)
                        winner_std.append(df_std.loc[index,columnIndex+'_std'])
                if len(winner)>1: #more than two winner, check std

                    min_std = min(winner_std)
                    winner_index=[i for i, x in enumerate(winner_std) if x == min_std]
                    winner_filter=np.array(winner)[winner_index]
                    filter=list(set(winner) - set(winner_filter))
                    for each in filter:
                        row[each]=False
            df_compare=df_compare.replace({10: np.nan})
            df_compare['winner'] = df.mul(df.columns.to_series()).apply(','.join, axis=1).str.strip(',')
            df_compare=df_compare[['species', 'antibiotics']+tool_list+['max_'+fscore]+[x+'_std' for x in tool_list]+['winner' ]]

            wb = load_workbook(path_table_results3_1)
            ew = pd.ExcelWriter(path_table_results3_1)
            ew.book = wb
            df_compare.to_excel(ew,sheet_name = (eachfold))
            ew.save()
        #--------------------------------
        #mean +- std verson


        df1 = pd.DataFrame(index=species_list)
        file_utility.make_dir(os.path.dirname(path_table_results3_2))
        df1.to_excel(path_table_results3_2, sheet_name='introduction')
        # foldset=['random folds','phylo-tree-based folds','KMA-based folds']
        # tool_list=['Point-/ResFinder', 'Aytan-Aktug', 'Seq2Geno2Pheno','PhenotypeSeeker', 'Kover']
            #each time count the cases the com_tool outperforms others.
        for eachfold in foldset:

            i=0
            for each_tool in tool_list:

                df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                for species, antibiotics_selected in zip(df_species, antibiotics):
                    species_sub=[species]
                    df_macro=combine_data(species_sub,level,fscore,[each_tool],[eachfold],output_path)
                    df_macro=df_macro.reset_index()
                    df_macro=df_macro.drop(columns=['index'])
                    df_macro[each_tool]=df_macro[fscore]
                    df_macro=df_macro[['species', 'antibiotics',each_tool]]
                    df_final= pd.concat([df_final,df_macro])

                if i==0:
                    df_compare=df_final
                else:
                    df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
                i+=1

            df_compare=df_compare[['species', 'antibiotics']+tool_list]


            with open('./data/AntiAcronym_dict.json') as f:
                map_acr = json.load(f)
            df_compare['antibiotics']=df_compare['antibiotics'].apply(lambda x: map_acr[x] )

            wb = load_workbook(path_table_results3_2)
            ew = pd.ExcelWriter(path_table_results3_2)
            ew.book = wb
            df_compare.to_excel(ew,sheet_name = (eachfold))
            ew.save()




    if step=='3':
        # paired T-test
        print('Now paired T-test')
        Presults={}
        for eachfold in foldset:
            i=0
            for each_tool in tool_list:

                df_final=pd.DataFrame(columns=['species', 'antibiotics', each_tool])
                for species in  species_list:

                    species_sub=[species]
                    df_score=combine_data(species_sub,level,fscore,[each_tool],[eachfold],output_path)
                    df_score=df_score.reset_index()
                    df_score=df_score.drop(columns=['index'])

                    df_score[fscore] = df_score[fscore].astype(str)
                    df_score[each_tool]=df_score[fscore].apply(lambda x:x.split('±')[0]) #get mean if there is mean
                    df_score[each_tool] = df_score[each_tool] .astype(float)
                    df_score=df_score[['species', 'antibiotics',each_tool]]
                    df_final= pd.concat([df_final,df_score])

                if i==0:
                    df_compare=df_final
                else:
                    df_compare=pd.merge(df_compare, df_final, how="left", on=['species', 'antibiotics'])
                i+=1


            df_mean=df_compare[tool_list]
            df_mean = df_mean.dropna()
            print('Paired T-test:')
            #T-test
            i_noDiff=0
            for each_com in list(itertools.combinations(tool_list, 2)):
                mean1 = df_mean[each_com[0]]
                mean2 = df_mean[each_com[1]]
                result=ttest_rel(mean1, mean2)
                pvalue = result[1]
                print(each_com,pvalue)
                Presults[str( [each_com]+[eachfold])]=pvalue
                if pvalue>=0.05:
                    i_noDiff+=1
        print('No. can not reject null hypothesis:',i_noDiff)
        with open(output_path+ 'Results/supplement_figures_tables/S6-3_software_Pvalue_'+fscore+'.json', 'w') as f:
            json.dump(Presults, f)
